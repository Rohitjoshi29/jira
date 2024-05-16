import requests
import openpyxl
from credentials import username, password

Sprint_id = input("Type your Sprint Id: ")
Board_nr = input("Type your Board number: ")

url = "https://planday.atlassian.net/rest/greenhopper/1.0/rapid/charts/sprintreport?rapidViewId=" + Board_nr + "&sprintId=" + Sprint_id

# Create authentication credentials
auth = (username, password)

response = requests.get(url, auth=auth)
data = response.json()

print (url)
# Check if the 'contents' key exists in the response
if 'contents' in data:
    contents = data['contents']

    # Check if the desired keys exist in the 'contents' section
    if 'completedIssuesInitialEstimateSum' in contents:
        completed_issues_initial_estimate_sum = contents['completedIssuesInitialEstimateSum']['value']
    else:
        completed_issues_initial_estimate_sum = None

    if 'completedIssuesEstimateSum' in contents:
        completed_issues_estimate_sum = contents['completedIssuesEstimateSum']['value']
    else:
        completed_issues_estimate_sum = None

    if 'allIssuesEstimateSum' in contents:
        all_issues_estimate_sum = contents['allIssuesEstimateSum']['value']
    else:
        all_issues_estimate_sum = None

    # Create or load the Excel file
    file_location = "C:/Python/Temp/velocity_chart.xlsx"
    wb = openpyxl.Workbook()

    # Select the active sheet
    sheet = wb.active

    # Write the headers
    sheet.cell(row=1, column=1, value='Sprint_id')
    sheet.cell(row=1, column=2, value='Board_nr')
    sheet.cell(row=1, column=3, value='completedIssuesInitialEstimateSum')
    sheet.cell(row=1, column=4, value='completedIssuesEstimateSum')
    sheet.cell(row=1, column=5, value='allIssuesEstimateSum')

    # Write the values
    sheet.cell(row=2, column=1, value=Sprint_id)
    sheet.cell(row=2, column=2, value=Board_nr)
    sheet.cell(row=2, column=3, value=completed_issues_initial_estimate_sum)
    sheet.cell(row=2, column=4, value=completed_issues_estimate_sum)
    sheet.cell(row=2, column=5, value=all_issues_estimate_sum)

    # Save the Excel file
    wb.save(file_location)
    print(f"Data saved to '{file_location}' successfully.")
else:
    print("'contents' key not found in the response")

import requests
import openpyxl
from credentials import username, password


def save_sprint_data(board_nr, sprint_ids):
    username = "your_username"
    password = "your_password"

    # Create or load the Excel file
    file_location = "C:/Python/Temp/velocity_chart.xlsx"
    wb = openpyxl.Workbook()

    # Select the active sheet
    sheet = wb.active

    # Write the headers
    sheet.cell(row=1, column=1, value='Sprint_id')
    sheet.cell(row=1, column=2, value='Board_nr')
    sheet.cell(row=1, column=3, value='completedIssuesInitialEstimateSum')
    sheet.cell(row=1, column=4, value='completedIssuesEstimateSum')
    sheet.cell(row=1, column=5, value='allIssuesEstimateSum')

    for sprint_id in sprint_ids:
        url = f"https://planday.atlassian.net/rest/greenhopper/1.0/rapid/charts/sprintreport?rapidViewId={board_nr}&sprintId={sprint_id}"

        # Create authentication credentials
        auth = (username, password)

        try:
            response = requests.get(url, auth=auth)
            data = response.json()

            print(url)
            print(data)  # Print the response JSON

            if 'contents' in data:
                contents = data['contents']
                # Rest of the code to process the data and save it to Excel

            else:
                print(f"'contents' key not found in the response for Sprint ID {sprint_id}")

        except Exception as e:
            print(f"An error occurred while processing Sprint ID {sprint_id}: {str(e)}")
            continue

    # Save the Excel file
    wb.save(file_location)
    print(f"Data saved to '{file_location}' successfully.")

# Pass parameters at runtime
board_nr = input("Type your Board number: ")
sprint_ids = input("Type your Sprint IDs (comma-separated): ").split(',')

save_sprint_data(board_nr, sprint_ids)
233